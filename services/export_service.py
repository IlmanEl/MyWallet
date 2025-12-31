import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from services.database import db
from utils.helpers import format_currency
import logging
import os

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to Excel"""

    @staticmethod
    def export_to_excel(user_id: int, start_date: datetime, end_date: datetime) -> str:
        """Export transactions to Excel file"""
        try:
            # Get transactions
            transactions = db.get_transactions(
                user_id=user_id,
                start_date=start_date,
                end_date=end_date,
                limit=10000
            )

            # Create workbook
            wb = openpyxl.Workbook()

            # Sheet 1: All Transactions
            ExportService._create_transactions_sheet(wb, transactions)

            # Sheet 2: Category Summary
            ExportService._create_category_summary_sheet(wb, transactions)

            # Sheet 3: Monthly Summary
            ExportService._create_monthly_summary_sheet(wb, user_id, start_date, end_date)

            # Save file
            filename = f"MyWallet_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(Config.TEMP_DIR, filename)

            wb.save(filepath)
            logger.info(f"Excel export created: {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"Error creating Excel export: {e}")
            raise

    @staticmethod
    def _create_transactions_sheet(wb, transactions):
        """Create sheet with all transactions"""
        ws = wb.active
        ws.title = "Все транзакции"

        # Headers
        headers = ["Дата", "Тип", "Категория", "Сумма (грн)", "Описание", "Способ оплаты", "Проект"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for t in transactions:
            ws.append([
                datetime.fromisoformat(t['date']).strftime('%d.%m.%Y %H:%M') if t.get('date') else '',
                'Доход' if t['type'] == 'income' else 'Расход',
                t.get('category', ''),
                t['amount'],
                t.get('description', ''),
                t.get('payment_method', ''),
                t.get('project', '')
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def _create_category_summary_sheet(wb, transactions):
        """Create sheet with category summary"""
        ws = wb.create_sheet("Статистика по категориям")

        # Calculate totals by category
        expense_totals = {}
        income_totals = {}

        for t in transactions:
            category = t.get('category', 'Без категории')
            if t['type'] == 'expense':
                expense_totals[category] = expense_totals.get(category, 0) + t['amount']
            else:
                income_totals[category] = income_totals.get(category, 0) + t['amount']

        # Headers
        ws.append(["РАСХОДЫ"])
        ws.append(["Категория", "Сумма (грн)", "Процент"])

        # Add expense data
        total_expense = sum(expense_totals.values())
        for category, amount in sorted(expense_totals.items(), key=lambda x: x[1], reverse=True):
            percentage = (amount / total_expense * 100) if total_expense > 0 else 0
            ws.append([category, amount, f"{percentage:.1f}%"])

        ws.append([])
        ws.append(["ИТОГО РАСХОДОВ", total_expense, "100%"])

        # Income section
        ws.append([])
        ws.append(["ДОХОДЫ"])
        ws.append(["Категория", "Сумма (грн)", "Процент"])

        total_income = sum(income_totals.values())
        for category, amount in sorted(income_totals.items(), key=lambda x: x[1], reverse=True):
            percentage = (amount / total_income * 100) if total_income > 0 else 0
            ws.append([category, amount, f"{percentage:.1f}%"])

        ws.append([])
        ws.append(["ИТОГО ДОХОДОВ", total_income, "100%"])

        # Balance
        ws.append([])
        ws.append(["БАЛАНС", total_income - total_expense])

    @staticmethod
    def _create_monthly_summary_sheet(wb, user_id, start_date, end_date):
        """Create sheet with monthly summary"""
        ws = wb.create_sheet("Месячная сводка")

        balance_data = db.get_balance(user_id)

        ws.append(["Месячная сводка"])
        ws.append([])
        ws.append(["Период:", f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"])
        ws.append([])
        ws.append(["Общий доход:", balance_data['total_income']])
        ws.append(["Общий расход:", balance_data['total_expense']])
        ws.append(["Баланс:", balance_data['balance']])


export_service = ExportService()
